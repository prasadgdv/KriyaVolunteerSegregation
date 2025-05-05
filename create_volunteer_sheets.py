import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import glob
import shutil

def create_volunteer_sheets(input_file_path):
    """
    Python implementation of the createVolunteerSheets function
    
    Args:
        input_file_path: Path to the Excel file to process
    """
    # First pre-process the input file to clean mobile numbers
    try:
        print("Pre-processing Excel file to clean mobile numbers...")
        # Read only first sheet for mobile number cleaning
        df = pd.read_excel(input_file_path, sheet_name=0)
        
        # Find the mobile column
        mobile_col = None
        for col in df.columns:
            if 'mobile' in str(col).lower() or 'phone' in str(col).lower():
                mobile_col = col
                break
                
        # If mobile column exists, clean the values
        modified = False
        if mobile_col:
            print(f"Found mobile column: '{mobile_col}'")
            
            # Count issues before fixing
            empty_before = df[mobile_col].isna().sum()
            error_before = sum(1 for val in df[mobile_col] if isinstance(val, str) and (val == "#ERROR!" or val.startswith("#")))
            
            # Fix empty/NaN values
            df[mobile_col] = df[mobile_col].fillna("1111111111")
            
            # Fix #ERROR! values and other problematic values
            for i, val in enumerate(df[mobile_col]):
                if pd.isna(val) or (isinstance(val, str) and (val == "#ERROR!" or val.startswith("#"))):
                    df.at[i, mobile_col] = "1111111111"
                    modified = True
                elif val == "":
                    df.at[i, mobile_col] = "1111111111"
                    modified = True
                # Try to clean mobile numbers if they contain invalid characters
                elif isinstance(val, str):
                    cleaned_val = ""
                    for char in val:
                        if char.isdigit() or char in ['+', '-', ' ', '(', ')']:
                            cleaned_val += char
                    if cleaned_val == "":
                        df.at[i, mobile_col] = "1111111111"
                        modified = True
                    elif cleaned_val != val:
                        df.at[i, mobile_col] = cleaned_val
                        modified = True
            
            # Count issues after fixing
            empty_after = df[mobile_col].isna().sum()
            error_after = sum(1 for val in df[mobile_col] if isinstance(val, str) and (val == "#ERROR!" or val.startswith("#")))
            
            print(f"Fixed {empty_before} empty values and {error_before} error values in mobile column")
            print(f"Remaining empty: {empty_after}, Remaining errors: {error_after}")
            
            if modified:
                # Read all sheets from the original file
                print("Reading all sheets from the Excel file...")
                excel_file = pd.ExcelFile(input_file_path)
                all_sheets = excel_file.sheet_names
                
                # Save to a temporary file with all sheets
                temp_file = input_file_path + ".temp.xlsx"
                with pd.ExcelWriter(temp_file) as writer:
                    # First sheet with cleaned data
                    df.to_excel(writer, sheet_name=all_sheets[0], index=False)
                    
                    # Copy rest of the sheets
                    if len(all_sheets) > 1:
                        for sheet_name in all_sheets[1:]:
                            sheet_df = pd.read_excel(input_file_path, sheet_name=sheet_name)
                            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"Saved cleaned data to temporary file with {len(all_sheets)} sheets")
                
                # Use the temporary file for the rest of the process
                input_file_path = temp_file
    except Exception as e:
        print(f"Error during pre-processing: {str(e)}")
        print("Continuing with original Excel file...")
    
    # Extract district name from the input file name
    input_file_name = os.path.basename(input_file_path)
    district_name = os.path.splitext(input_file_name)[0]
    if district_name.endswith('.temp'):
        district_name = district_name[:-5]  # Remove .temp suffix if present
    if ' D' in district_name:  # If format is like "Kurnool D.xlsx"
        district_name = district_name.split(' D')[0]
    
    print(f"Extracted district name: {district_name}")
    
    # Create excels_[district] folder
    base_dir = os.path.dirname(os.path.abspath(input_file_path))
    excels_folder = os.path.join(base_dir, f"excels_{district_name.lower()}")
    if not os.path.exists(excels_folder):
        os.makedirs(excels_folder)
        print(f"Created folder: {excels_folder}")
    
    # Load the workbook to get sheet/tab names
    wb = openpyxl.load_workbook(input_file_path)
    sheets = wb.sheetnames
    
    # Check if workbook has multiple sheets
    if len(sheets) <= 1:
        print(f"Processing single sheet: {wb.active.title}")
        sheet = wb.active
        process_sheet(sheet, input_file_path, district_name, excels_folder)
    else:
        print(f"Found {len(sheets)} tabs in the workbook: {', '.join(sheets)}")
        total_volunteers = 0
        
        # Process each sheet
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            process_count = process_sheet(sheet, input_file_path, district_name, excels_folder, sheet_name)
            if process_count:
                total_volunteers += process_count
        
        print(f"\nTotal volunteers processed across all tabs: {total_volunteers}")
    
    # Create corresponding PDF folders structure
    pdf_base_folder = os.path.join(base_dir, f"pdfs_{district_name.lower()}")
    if not os.path.exists(pdf_base_folder):
        os.makedirs(pdf_base_folder)
        print(f"Created PDF base folder: {pdf_base_folder}")
    
    # Create PDF tab subfolders (matching Excel folders)
    for tab_folder in os.listdir(excels_folder):
        tab_path = os.path.join(excels_folder, tab_folder)
        if os.path.isdir(tab_path):
            pdf_tab_folder = os.path.join(pdf_base_folder, tab_folder)
            if not os.path.exists(pdf_tab_folder):
                os.makedirs(pdf_tab_folder)
                print(f"Created PDF tab folder: {pdf_tab_folder}")
    
    # Clean up temporary file if it exists
    if input_file_path.endswith('.temp.xlsx') and os.path.exists(input_file_path):
        try:
            os.remove(input_file_path)
            print(f"Cleaned up temporary file")
        except Exception as e:
            print(f"Could not remove temporary file: {str(e)}")
            
    return True  # Return True to indicate success

def process_sheet(sheet, input_file_path, district_name, excels_folder, sheet_name=None):
    """
    Process a single sheet/tab from the Excel file
    
    Args:
        sheet: The worksheet object to process
        input_file_path: Path to the Excel file
        district_name: Name of the district extracted from file name
        excels_folder: Base folder for Excel files
        sheet_name: Name of the sheet being processed (if None, use sheet.title)
    
    Returns:
        int: Number of volunteers processed
    """
    # Use sheet title if sheet_name is not provided
    if sheet_name is None:
        sheet_name = sheet.title
        
    print(f"\nProcessing sheet: {sheet_name}")
    
    # Get all data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    
    # Track volunteer records by name and phone number
    volunteer_records = {}  # Key: (volunteer_name, phone_number), Value: list of records
    
    # Step 1: Group all records by volunteer name and phone number
    for i in range(1, len(data)):  # Start from row 1 (skipping header)
        if i >= len(data) or len(data[i]) < 9:  # Skip invalid rows (need at least 9 columns)
            continue
            
        volunteer_name = data[i][7]  # volunteerName is in column 8 (index 7)
        if not volunteer_name:  # Skip rows where volunteer name is empty
            continue
            
        # Get phone for this volunteer at index 8 (column 9)
        volunteer_phone = data[i][8] if len(data[i]) > 8 and data[i][8] is not None else ""
        if isinstance(volunteer_phone, (int, float)):
            volunteer_phone = str(int(volunteer_phone))
        
        # Create a unique key based on name + phone
        key = (volunteer_name, volunteer_phone)
        
        # Group by name and phone
        if key not in volunteer_records:
            volunteer_records[key] = []
        volunteer_records[key].append(data[i])
    
    # Count how many distinct volunteers we have after grouping
    unique_volunteers = len(volunteer_records)
    
    # Count duplicate names and create a mapping of volunteer names to their count
    volunteer_name_counts = {}
    for (name, _) in volunteer_records.keys():
        if name not in volunteer_name_counts:
            volunteer_name_counts[name] = 0
        volunteer_name_counts[name] += 1
    
    duplicate_names = sum(1 for count in volunteer_name_counts.values() if count > 1)
    print(f"Found {unique_volunteers} unique volunteers in '{sheet_name}' tab")
    if duplicate_names > 0:
        print(f"Detected {duplicate_names} volunteer names with multiple phone numbers")
    
    # Create tab folder (using sheet name)
    safe_tab_name = str(sheet_name).replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
    tab_folder = os.path.join(excels_folder, safe_tab_name.lower())
    if not os.path.exists(tab_folder):
        os.makedirs(tab_folder)
        print(f"Created tab folder: {tab_folder}")
    
    # Process all volunteers
    processed_count = 0
    
    # Sort volunteer keys (name, phone) alphabetically by name first
    sorted_keys = sorted(volunteer_records.keys(), key=lambda k: k[0])
    
    for key in sorted_keys:
        volunteer_name, volunteer_phone = key
        volunteer_data = volunteer_records[key]
        
        # Create a new workbook for this volunteer
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # Add header information in a merged cell
        new_sheet.merge_cells('A1:F1')
        header_cell = new_sheet['A1']
        
        # Get volunteer number from first record
        volunteer_number = volunteer_data[0][8]
        if isinstance(volunteer_number, (int, float)):
            volunteer_number = str(int(volunteer_number))  # Convert to int first to remove decimal, then to string
        elif volunteer_number is None:  # Handle None values
            volunteer_number = ""
        
        header_cell.value = f"Kriya VolunteerName: {volunteer_name}    Volunteer number: {volunteer_number}"
        
        # Style for header
        header_cell.font = Font(name='Verdana', size=11, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='DF6666', end_color='DF6666', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center')
        
        # Apply border to all cells in the merged header region
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for col in range(1, 7):  # A1 through F1
            cell = new_sheet.cell(row=1, column=col)
            cell.border = thin_border
        
        # Add column headers
        headers = ["S No", "Mandal", "JSP Id", "Name", "Mobile", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = new_sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(name='Verdana', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='DF6666', end_color='DF6666', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Add records for the volunteer
        for j, vol_data in enumerate(volunteer_data):
            # S No, then columns 2-6 from data, plus empty status
            new_sheet.cell(row=j+3, column=1).value = j + 1
            
            for k in range(4):  # Columns 2-5 (Mandal, JSP Id, Name, Mobile)
                col_idx = k + 2  # Column index in sheet (2-5)
                data_idx = k + 2  # Index in vol_data (start at 2 for Mandal)
                
                cell_value = vol_data[data_idx] if data_idx < len(vol_data) else ""
                
                # Format phone numbers (typically in column 5, which is Mobile)
                if col_idx == 5 and isinstance(cell_value, (int, float)):
                    cell_value = str(int(cell_value))
                
                new_sheet.cell(row=j+3, column=col_idx).value = cell_value
            
            # Empty status cell
            new_sheet.cell(row=j+3, column=6).value = ""
        
        # Set column widths
        column_widths = [50, 150, 100, 200, 100, 100]
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            new_sheet.column_dimensions[column_letter].width = width / 7  # Convert approximate Google Sheets width to Excel
        
        # Apply borders to all cells
        last_row = len(volunteer_data) + 2
        for row in new_sheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
        
        # Set font and alignment for data rows
        for row in new_sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=6):
            for cell in row:
                cell.font = Font(name='Verdana', size=11)
                if cell.column == 4:  # Align Name column to left
                    cell.alignment = Alignment(horizontal='left')
                elif cell.column == 2:  # Align Mandal column to left
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')
        
        # Create filename - if duplicate volunteer name, append phone number
        safe_volunteer_name = volunteer_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
        
        if volunteer_name_counts[volunteer_name] > 1:
            # Append phone number to filename if this volunteer name appears multiple times with different phones
            safe_filename = f"{safe_volunteer_name}_{volunteer_phone}"
        else:
            # No duplicate names, use just the volunteer name
            safe_filename = safe_volunteer_name
        
        # Save the Excel file
        excel_path = os.path.join(tab_folder, f"{safe_filename}.xlsx")
        try:
            new_wb.save(excel_path)
            # Increment processed count
            processed_count += 1
            
            # Show phone number in display if this is a duplicate name case
            if volunteer_name_counts[volunteer_name] > 1:
                display_name = f"{volunteer_name} (Phone: {volunteer_phone})"
            else:
                display_name = volunteer_name
                
            print(f"Created Excel for volunteer {processed_count}/{unique_volunteers}: {display_name}")
        except Exception as e:
            print(f"Error saving file for {volunteer_name}: {e}")
    
    print(f"\nProcessed {processed_count} volunteers for tab '{sheet_name}'. Files saved in {tab_folder}")
    return processed_count

def repair_excel_file(excel_file_path, volunteer_data):
    """
    Repair a corrupted Excel file using data extracted from the master file
    
    Args:
        excel_file_path (str): Path to the Excel file to repair
        volunteer_data (list): List of data rows for the volunteer
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Extract volunteer name from filename
        volunteer_name = os.path.splitext(os.path.basename(excel_file_path))[0]
        print(f"Repairing Excel file for: {volunteer_name}")
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Add header information in a merged cell
        sheet.merge_cells('A1:F1')
        header_cell = sheet['A1']
        
        # Extract volunteer number from data
        volunteer_number = volunteer_data[0][8] if len(volunteer_data) > 0 and len(volunteer_data[0]) > 8 else ""
        if isinstance(volunteer_number, (int, float)):
            volunteer_number = str(int(volunteer_number))
        
        header_cell.value = f"Kriya VolunteerName: {volunteer_name}    Volunteer number: {volunteer_number}"
        
        # Style for header
        header_cell.font = Font(name='Verdana', size=11, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='DF6666', end_color='DF6666', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center')
        
        # Apply border to all cells in the merged header region
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for col in range(1, 7):  # A1 through F1
            cell = sheet.cell(row=1, column=col)
            cell.border = thin_border
        
        # Add column headers
        headers = ["S No", "Mandal", "JSP Id", "Name", "Mobile", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(name='Verdana', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='DF6666', end_color='DF6666', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Add records for the volunteer
        for j, vol_data in enumerate(volunteer_data):
            # S No
            sheet.cell(row=j+3, column=1).value = j + 1
            
            # Copy data from volunteer_data (Mandal, JSP Id, Name, Mobile)
            for k in range(4):
                col_idx = k + 2  # Column index in sheet (2-5)
                data_idx = k + 2  # Index in vol_data (start at 2 for Mandal)
                
                cell_value = vol_data[data_idx] if data_idx < len(vol_data) else ""
                
                # Format phone numbers
                if col_idx == 5 and isinstance(cell_value, (int, float)):
                    cell_value = str(int(cell_value))
                
                sheet.cell(row=j+3, column=col_idx).value = cell_value
            
            # Empty status cell
            sheet.cell(row=j+3, column=6).value = ""
        
        # Set column widths
        column_widths = [50, 150, 100, 200, 100, 100]
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width / 7
        
        # Apply borders to all cells
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        last_row = len(volunteer_data) + 2
        for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
        
        # Set font and alignment for data rows
        for row in sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=6):
            for cell in row:
                cell.font = Font(name='Verdana', size=11)
                if cell.column == 4:  # Align Name column to left
                    cell.alignment = Alignment(horizontal='left')
                elif cell.column == 2:  # Align Mandal column to left
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')
        
        # Create a backup of the original file if it exists
        if os.path.exists(excel_file_path):
            backup_path = excel_file_path + ".bak"
            shutil.copy2(excel_file_path, backup_path)
            print(f"Created backup of original file: {os.path.basename(backup_path)}")
        
        # Save the repaired Excel file
        wb.save(excel_file_path)
        print(f"Saved repaired Excel file: {os.path.basename(excel_file_path)}")
        
        return True
        
    except Exception as e:
        print(f"Error repairing Excel file: {str(e)}")
        return False

# Function to select an Excel file from a list
def select_excel_file():
    print("Searching for Excel files in the current directory...")
    
    # Get the script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Find all Excel files in the current directory only (not recursive)
    excel_files = []
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(glob.glob(os.path.join(script_dir, ext)))
    
    # Filter out temporary files and processed files
    excel_files = [f for f in excel_files if not os.path.basename(f).endswith(('_processed.xlsx', '.temp.xlsx', '.temp_processed.xlsx'))]
    
    # Sort the files alphabetically
    excel_files = sorted(list(set(excel_files)))
    
    if not excel_files:
        print("No Excel files found in the current directory.")
        user_input = input("Please enter the full path to an Excel file: ")
        return user_input
    
    print("\nExcel files in the current directory:")
    
    # Print files with their index
    file_map = {}
    for idx, file_path in enumerate(excel_files, start=1):
        file_name = os.path.basename(file_path)
        file_map[idx] = file_path
        print(f"[{idx}] {file_name}")
    
    while True:
        try:
            choice = input("\nEnter the number of the file you want to process (or 'q' to quit): ")
            
            if choice.lower() == 'q':
                return None
                
            choice = int(choice)
            if choice in file_map:
                return file_map[choice]
            else:
                print(f"Invalid selection. Please enter a number between 1 and {len(file_map)}.")
        except ValueError:
            print("Please enter a valid number or 'q' to quit.")

if __name__ == "__main__":
    print("=" * 50)
    print("Excel Volunteer Sheet Creator")
    print("=" * 50)
    
    # Get the Excel file path from the user
    excel_file = select_excel_file()
    
    if excel_file:
        print(f"\nProcessing file: {excel_file}")
        success = create_volunteer_sheets(excel_file)
        if success:
            print("\nVolunteer sheets created successfully!")
            
            # Ask if user wants to convert to PDF
            convert_to_pdf = input("\nDo you want to convert Excel files to PDF? (y/n): ").lower() == 'y'
            if convert_to_pdf:
                try:
                    import subprocess
                    from convert_to_pdf import process_excel_folders
                    print("\nStarting PDF conversion process...")
                    process_excel_folders()
                except Exception as e:
                    print(f"Error during PDF conversion: {str(e)}")
                    print("You can run convert_to_pdf.py separately to convert files.")
            else:
                print("\nExcel files created. You can run convert_to_pdf.py later to convert them to PDF.")
        else:
            print("\nFailed to create volunteer sheets. Please check the file format.")
    else:
        print("Operation cancelled.")