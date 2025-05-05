import os
import sys
import glob
import time
import argparse
from pathlib import Path
import concurrent.futures
import threading
import math

# Thread-safe counter for progress tracking
class AtomicCounter:
    def __init__(self, initial=0):
        self.value = initial
        self._lock = threading.Lock()
        
    def increment(self):
        with self._lock:
            self.value += 1
            return self.value
            
    def get(self):
        with self._lock:
            return self.value

def convert_excel_file_to_pdf(excel_file, output_folder, total_files, counter, excel_app=None):
    """
    Convert a single Excel file to PDF
    
    Args:
        excel_file (str): Path to Excel file
        output_folder (str): Path to save PDF file
        total_files (int): Total number of files for progress reporting
        counter (AtomicCounter): Thread-safe counter for progress
        excel_app: Excel application instance (if None, will create a new one)
        
    Returns:
        tuple: (success, filename, error_message)
    """
    filename = os.path.basename(excel_file)
    pdf_name = os.path.splitext(filename)[0] + ".pdf"
    pdf_path = os.path.join(output_folder, pdf_name)
    
    # Get current progress count
    current = counter.increment()
    print(f"[{current}/{total_files}] Converting: {filename}...")
    
    # Create a new Excel application instance if one wasn't provided
    local_excel_app = None
    locally_created = False
    
    try:
        import win32com.client
        from win32com.client import constants
        
        if excel_app is None:
            local_excel_app = win32com.client.Dispatch("Excel.Application")
            local_excel_app.Visible = False
            local_excel_app.DisplayAlerts = False
            locally_created = True
        else:
            local_excel_app = excel_app
        
        # Open the workbook
        wb = local_excel_app.Workbooks.Open(os.path.abspath(excel_file))
        
        # Set print area to entire used range for all sheets
        for sheet in wb.Worksheets:
            # Get row count to calculate optimal settings
            used_range = sheet.UsedRange
            last_row = used_range.Row + used_range.Rows.Count - 1
            
            # Set page orientation to PORTRAIT mode
            sheet.PageSetup.Orientation = 1  # xlPortrait (1=Portrait, 2=Landscape)
            
            # Set paper size to A4
            sheet.PageSetup.PaperSize = 9  # xlPaperA4
            
            # OPTIMIZED SETTINGS FOR FITTING MORE ROWS:
            
            # 1. Set margins with significantly increased top margin for better spacing
            sheet.PageSetup.LeftMargin = 7.2    # 0.1 inches (left margin remains small)
            sheet.PageSetup.RightMargin = 7.2   # 0.1 inches (right margin remains small)
            sheet.PageSetup.TopMargin = 43.2    # 0.6 inches (increased significantly for more top space)
            sheet.PageSetup.BottomMargin = 21.6 # 0.3 inches
            sheet.PageSetup.HeaderMargin = 7.2  # 0.1 inches
            sheet.PageSetup.FooterMargin = 7.2  # 0.1 inches
            
            # 2. Handle scaling to fit more rows
            sheet.PageSetup.Zoom = False  # Don't use percentage zoom
            sheet.PageSetup.FitToPagesWide = 1  # Fit width on one page
            
            # Force fitting to pages tall if needed
            if last_row <= 45:  # For sheets with fewer rows, use automatic scaling
                sheet.PageSetup.FitToPagesTall = 1  # Fit all rows on one page
            else:
                # Let it flow to multiple pages with better row density
                sheet.PageSetup.FitToPagesTall = False
                
                # Calculate approximate rows per page (with minimal margins, around 45-50 per page)
                rows_per_page = 48  # Optimized target for A4 landscape with minimal margins
                pages_needed = (last_row + 2) / rows_per_page  # +2 for headers
                pages_needed = max(1, round(pages_needed))
                
                # Adjust scaling for best fit
                if pages_needed > 1:
                    sheet.PageSetup.FitToPagesTall = pages_needed
            
            # 3. Center content horizontally only
            sheet.PageSetup.CenterHorizontally = True
            sheet.PageSetup.CenterVertically = False
            
            # 4. Set print titles to repeat header rows
            sheet.PageSetup.PrintTitleRows = "$1:$2"
            
            # 5. Turn off gridlines and row/column headings
            sheet.PageSetup.PrintGridlines = False
            sheet.PageSetup.PrintHeadings = False
        
        # Export to PDF with optimized settings
        wb.ExportAsFixedFormat(
            Type=0,  # PDF format
            Filename=os.path.abspath(pdf_path),
            Quality=0,  # Standard quality
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
        
        # Close the workbook without saving changes
        wb.Close(SaveChanges=False)
        
        print(f"✓ Successfully created: {pdf_name}")
        return (True, filename, None)
        
    except Exception as e:
        error_msg = str(e)
        print(f"✗ Error converting {filename}: {error_msg}")
        
        # Try to close the workbook if it's still open
        try:
            if 'wb' in locals():
                wb.Close(SaveChanges=False)
        except:
            pass
            
        return (False, filename, error_msg)
        
    finally:
        # Clean up: quit the locally created Excel instance
        if locally_created and local_excel_app is not None:
            try:
                local_excel_app.Quit()
            except:
                pass

def convert_excel_to_pdf(input_folder, output_folder, test_mode=False, max_workers=None):
    """
    Convert Excel files to PDF with adjusted settings for fitting more rows per page.
    Uses parallel processing to speed up conversion.
    
    Args:
        input_folder (str): Path to folder containing Excel files
        output_folder (str): Path to save PDF files
        test_mode (bool): If True, only convert first 2 files for testing
        max_workers (int): Maximum number of parallel workers (None=auto)
        
    Returns:
        tuple: (success_count, error_count, failed_files)
    """
    print(f"Excel to PDF Conversion - Optimized for 45+ rows per page (WITH PARALLEL PROCESSING)")
    print(f"=======================================================================")
    print(f"Input folder: {input_folder}")
    print(f"Output folder: {output_folder}")
    
    # Determine optimal number of workers
    if max_workers is None:
        # Use number of cores, but cap at 4 to prevent too many Excel instances
        max_workers = min(4, os.cpu_count() or 2)
    
    print(f"Using {max_workers} parallel workers for conversion")
    
    if test_mode:
        print("TESTING MODE: Only converting first 2 files")
    
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created output folder: {output_folder}")
    
    # Find all Excel files in the input folder
    excel_files = glob.glob(os.path.join(input_folder, "*.xlsx"))
    
    if not excel_files:
        print(f"No Excel files found in {input_folder}")
        return 0, 0, []
    
    if test_mode:
        # Limit to first 2 files for testing
        excel_files = excel_files[:2]
        print(f"Testing with {len(excel_files)} Excel files")
    else:
        print(f"Found {len(excel_files)} Excel files to convert")
    
    # Import the win32com module for Excel automation
    try:
        import win32com.client
        from win32com.client import constants
    except ImportError:
        print("ERROR: Required module 'pywin32' not found.")
        print("Please run: pip install pywin32")
        return 0, 0, []
    
    # Initialize counter for tracking progress
    counter = AtomicCounter(0)
    
    # Convert files in parallel
    success_count = 0
    error_count = 0
    failed_files = []
    
    # Determine whether to use single-process or multi-process approach
    if len(excel_files) <= 3 or max_workers == 1:
        # For small batches, use a single Excel instance for all files
        print("Using single-process mode for small batch...")
        try:
            # Initialize the Excel application
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            for excel_file in excel_files:
                result = convert_excel_file_to_pdf(excel_file, output_folder, len(excel_files), counter, excel)
                if result[0]:
                    success_count += 1
                else:
                    error_count += 1
                    failed_files.append((result[1], result[2]))
                    
        except Exception as e:
            print(f"Error in single-process conversion: {str(e)}")
        finally:
            # Clean up: quit Excel
            try:
                if 'excel' in locals():
                    excel.Quit()
            except:
                pass
    else:
        # For larger batches, use parallel processing
        print(f"Using multi-process mode with {max_workers} workers...")
        
        # Calculate batch size for better performance (to avoid creating/destroying too many Excel instances)
        batch_size = max(3, math.ceil(len(excel_files) / max_workers))
        batches = [excel_files[i:i + batch_size] for i in range(0, len(excel_files), batch_size)]
        
        print(f"Divided {len(excel_files)} files into {len(batches)} batches of ~{batch_size} files each")
        
        # Process batches in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit each batch for processing
            batch_futures = []
            for batch in batches:
                future = executor.submit(process_batch, batch, output_folder, len(excel_files), counter)
                batch_futures.append(future)
            
            # Collect results
            for future in concurrent.futures.as_completed(batch_futures):
                batch_success, batch_error, batch_failed = future.result()
                success_count += batch_success
                error_count += batch_error
                failed_files.extend(batch_failed)
    
    print(f"\nConversion completed: {success_count} successful, {error_count} failed")
    print(f"PDF files saved in: {output_folder}")
    
    if test_mode:
        print("\nTEST RUN COMPLETED - Please check the generated PDFs to verify row count and formatting.")
        print("If the results look good, run the script again without test mode to convert all files.")
    
    return success_count, error_count, failed_files

def process_batch(files_batch, output_folder, total_files, counter):
    """
    Process a batch of Excel files with a single Excel instance
    
    Args:
        files_batch (list): List of Excel files to process
        output_folder (str): Path to save PDF files
        total_files (int): Total number of files for progress reporting
        counter (AtomicCounter): Thread-safe counter for progress
        
    Returns:
        tuple: (success_count, error_count, failed_files)
    """
    success_count = 0
    error_count = 0
    failed_files = []
    
    # Initialize COM for this thread
    try:
        import win32com.client
        import pythoncom
        
        # Initialize COM for this thread - THIS IS THE KEY FIX
        pythoncom.CoInitialize()
        
        # Create a single Excel instance for this batch
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Process each file in the batch
        for excel_file in files_batch:
            result = convert_excel_file_to_pdf(excel_file, output_folder, total_files, counter, excel)
            if result[0]:
                success_count += 1
            else:
                error_count += 1
                failed_files.append((result[1], result[2]))
    
    except Exception as e:
        print(f"Error processing batch: {str(e)}")
    finally:
        # Clean up: quit Excel and uninitialize COM
        try:
            if 'excel' in locals():
                excel.Quit()
            pythoncom.CoUninitialize()  # Clean up COM for this thread
        except:
            pass
    
    return success_count, error_count, failed_files

def create_failed_list_excel(output_folder, failed_files, tab_name):
    """
    Create an Excel file with the list of files that failed during conversion.
    
    Args:
        output_folder (str): Path to the PDF output folder
        failed_files (list): List of tuples (filename, error_message)
        tab_name (str): Name of the tab being processed
    """
    if not failed_files:
        print("No failed files to log.")
        return
    
    try:
        # Import openpyxl for creating the Excel file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("WARNING: Could not create failed list (openpyxl module not found).")
        return
    
    # Create a new workbook for the failed list
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Failed Files"
    
    # Add header
    sheet["A1"] = f"Failed PDF Conversions - {tab_name}"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells("A1:C1")
    
    # Add column headers
    sheet["A3"] = "File Name"
    sheet["B3"] = "Error Message"
    sheet["C3"] = "Date/Time"
    
    # Style headers
    for cell in sheet["3:3"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
    for i, (filename, error) in enumerate(failed_files, start=4):
        sheet[f"A{i}"] = filename
        sheet[f"B{i}"] = error
        sheet[f"C{i}"] = current_time
    
    # Auto-adjust column widths
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 60
    sheet.column_dimensions["C"].width = 20
    
    # Save the failed list Excel file
    failed_list_path = os.path.join(output_folder, f"failed_list_{tab_name}.xlsx")
    try:
        wb.save(failed_list_path)
        print(f"Created failed files list: {failed_list_path}")
    except Exception as e:
        print(f"Error saving failed files list: {str(e)}")

def find_district_excel_folders():
    """
    Find all district excel folders (e.g., excels_kurnool, excels_wg) in the workspace
    
    Returns:
        list: List of district folders found
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    district_folders = []
    
    # Look for folders starting with "excels_"
    for item in os.listdir(base_dir):
        if item.startswith("excels_") and os.path.isdir(os.path.join(base_dir, item)):
            district_folders.append(item)
    
    return district_folders

def process_excel_folders():
    """
    Process all Excel subfolders in the district folders (excels_kurnool, excels_wg, etc.)
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Find all district excel folders
    district_folders = find_district_excel_folders()
    
    if not district_folders:
        print(f"No district folders (excels_*) found. Please run create_volunteer_sheets.py first.")
        return
    
    print(f"Found {len(district_folders)} district folders to process:")
    for i, folder in enumerate(district_folders, 1):
        district_name = folder.replace("excels_", "")
        print(f"  {i}. {district_name}")
    
    # Ask user which district to process
    while True:
        try:
            choice = input("\nEnter district number to process (or 'all' for all districts): ")
            
            if choice.lower() == 'all':
                selected_districts = district_folders
                break
            else:
                idx = int(choice) - 1
                if 0 <= idx < len(district_folders):
                    selected_districts = [district_folders[idx]]
                    break
                else:
                    print(f"Invalid selection. Please enter a number between 1 and {len(district_folders)} or 'all'.")
        except ValueError:
            print("Please enter a valid number or 'all'.")
    
    # Process each selected district
    for district_folder in selected_districts:
        district_name = district_folder.replace("excels_", "")
        print(f"\nProcessing district: {district_name.upper()}")
        
        excels_folder = os.path.join(base_dir, district_folder)
        pdfs_folder = os.path.join(base_dir, f"pdfs_{district_name}")
        
        # Create PDF base folder if it doesn't exist
        if not os.path.exists(pdfs_folder):
            os.makedirs(pdfs_folder)
            print(f"Created PDF folder: {pdfs_folder}")
        
        # Get all mandal subfolders in the excels folder
        mandal_folders = [f for f in os.listdir(excels_folder) if os.path.isdir(os.path.join(excels_folder, f))]
        
        if not mandal_folders:
            print(f"No mandal subfolders found in: {excels_folder}")
            continue
        
        print(f"Found {len(mandal_folders)} mandal folders to process in {district_name}:")
        for i, mandal in enumerate(mandal_folders, 1):
            print(f"  {i}. {mandal}")
        
        # Ask user which mandals to process
        while True:
            try:
                choice = input(f"\nEnter mandal number(s) to process (comma-separated) or 'all' for all mandals in {district_name}: ")
                
                if choice.lower() == 'all':
                    selected_mandals = mandal_folders
                    break
                else:
                    indices = [int(idx.strip()) - 1 for idx in choice.split(',')]
                    selected_mandals = [mandal_folders[idx] for idx in indices if 0 <= idx < len(mandal_folders)]
                    if selected_mandals:
                        break
                    else:
                        print(f"Invalid selection. Please enter valid numbers between 1 and {len(mandal_folders)}")
            except ValueError:
                print("Please enter valid number(s) or 'all'.")
        
        # Process each selected mandal
        total_success = 0
        total_failed = 0
        
        for i, mandal in enumerate(selected_mandals, 1):
            print(f"\n[{i}/{len(selected_mandals)}] Processing mandal: {mandal}")
            
            # Setup input and output folders
            input_folder = os.path.join(excels_folder, mandal)
            output_folder = os.path.join(pdfs_folder, mandal)
            
            # Create output folder if it doesn't exist
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
                print(f"Created output folder: {output_folder}")
            
            # Convert Excel files to PDF
            success, failed, failed_files = convert_excel_to_pdf(input_folder, output_folder)
            total_success += success
            total_failed += failed
            
            # Create a failed list Excel file if there are any failed files
            if failed_files:
                create_failed_list_excel(output_folder, failed_files, mandal)
        
        print(f"\nDistrict {district_name} processing completed!")
        print(f"Total conversions: {total_success + total_failed}")
        print(f"Total successful: {total_success}")
        print(f"Total failed: {total_failed}")
        
        if total_failed > 0:
            print("\nNOTE: Failed file lists have been created in each mandal's PDF folder")
            print("Check the 'failed_list_*.xlsx' files for details on conversion failures.")
    
    print("\nAll selected districts have been processed!")

if __name__ == "__main__":
    # Parse command line arguments if provided
    parser = argparse.ArgumentParser(description="Convert Excel files to PDF with formatting optimized for more rows")
    parser.add_argument("--test", action="store_true", help="Test mode: process only the first 2 files")
    parser.add_argument("--manual", action="store_true", help="Manual mode: provide specific input/output folders")
    args = parser.parse_args()
    
    if args.manual:
        # Manual mode - ask for specific folders
        input_folder = input("Enter the path to the folder containing Excel files: ")
        output_folder = input("Enter the path to save PDF files: ")
        test_mode = args.test
        
        if not os.path.exists(input_folder):
            print(f"Error: Input folder '{input_folder}' does not exist.")
        else:
            success, failed, failed_files = convert_excel_to_pdf(input_folder, output_folder, test_mode)
            if failed_files:
                tab_name = os.path.basename(input_folder)
                create_failed_list_excel(output_folder, failed_files, tab_name)
                print(f"\nCreated failed files list in {output_folder}")
    else:
        # Automatic mode - process district folders
        print("Processing district folders...")
        process_excel_folders()
    
    print("\nPress Enter to exit...")
    input()